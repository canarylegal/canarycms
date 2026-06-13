"""Authenticated reporting endpoints (JSON + Excel)."""

from __future__ import annotations

from io import BytesIO
from typing import Annotated, Literal

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.accountant_pack_service import build_accountant_pack, preview_accountant_pack
from app.db import get_db
from app.deps import get_current_user
from app.models import CaseStatus, User
from app.reports_service import (
    enforce_fee_earner_ids,
    list_fee_earner_pick_users,
    pence_to_pounds_str,
    report_aged_debt,
    report_billing,
    report_cases,
    report_cases_opened,
    report_client_office_balances,
    report_events,
    report_exceptions,
    report_ledger_activity,
)
from app.schemas import (
    AccountantPackIn,
    AccountantPackPreviewOut,
    AccountantPackSectionOut,
    AgedDebtReportIn,
    BillingReportIn,
    CasesOpenedReportIn,
    CasesReportIn,
    EventsReportIn,
    ExceptionsReportIn,
    FeeEarnerPickOut,
    LedgerActivityReportIn,
    ReportFeeEarnerIdsIn,
)

router = APIRouter(prefix="/reports", tags=["reports"])

ReportFormat = Literal["json", "xlsx"]


def _wb_response(wb: Workbook, download_name: str) -> StreamingResponse:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


def _invoice_status_label(s: str) -> str:
    if s == "pending_approval":
        return "Pending approval"
    if s == "approved":
        return "Approved"
    if s == "voided":
        return "Voided"
    return s


@router.get("/fee-earners", response_model=list[FeeEarnerPickOut])
def fee_earner_options(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> list[FeeEarnerPickOut]:
    rows = list_fee_earner_pick_users(user, db)
    return [FeeEarnerPickOut(id=u.id, display_name=u.display_name, email=u.email) for u in rows]


@router.post("/client-office-balances")
def post_client_office_balances(
    body: ReportFeeEarnerIdsIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    rows, totals = report_client_office_balances(ids, db)
    if format == "json":
        return {
            "rows": [
                {
                    "case_id": str(r.case_id),
                    "case_number": r.case_number,
                    "client_name": r.client_name,
                    "matter_description": r.matter_description,
                    "fee_earner_name": r.fee_earner_name,
                    "client_balance_pence": r.client_balance_pence,
                    "office_balance_pence": r.office_balance_pence,
                }
                for r in rows
            ],
            "totals": totals,
        }
    wb = Workbook()
    ws = wb.active
    ws.title = "Client office balances"
    ws.append(
        [
            "Reference",
            "Client",
            "Matter description",
            "Fee earner",
            "Client balance (£)",
            "Office balance (£)",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.fee_earner_name,
                float(pence_to_pounds_str(r.client_balance_pence)),
                float(pence_to_pounds_str(r.office_balance_pence)),
            ]
        )
    ws.append(
        [
            "",
            "",
            "",
            "Total",
            float(pence_to_pounds_str(totals["client_balance_pence"])),
            float(pence_to_pounds_str(totals["office_balance_pence"])),
        ]
    )
    return _wb_response(wb, "canary-report-client-office-balances.xlsx")


@router.post("/billing")
def post_billing_report(
    body: BillingReportIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    rows, totals = report_billing(ids, db, date_from=body.date_from, date_to=body.date_to)
    if format == "json":
        return {
            "rows": [
                {
                    "invoice_id": str(r.invoice_id),
                    "case_number": r.case_number,
                    "client_name": r.client_name,
                    "invoice_number": r.invoice_number,
                    "invoice_status": r.invoice_status,
                    "invoice_status_label": _invoice_status_label(r.invoice_status),
                    "fee_earner_name": r.fee_earner_name,
                    "created_at": r.created_at.isoformat(),
                    "fees_ex_vat_pence": r.fees_ex_vat_pence,
                    "vat_pence": r.vat_pence,
                    "disbursements_ex_vat_pence": r.disbursements_ex_vat_pence,
                }
                for r in rows
            ],
            "totals": totals,
        }
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing"
    ws.append(
        [
            "Reference",
            "Client",
            "Invoice",
            "Status",
            "Fee earner",
            "Created (UTC)",
            "Fees ex VAT (£)",
            "VAT (£)",
            "Disbursements ex VAT (£)",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.case_number,
                r.client_name or "",
                r.invoice_number,
                _invoice_status_label(r.invoice_status),
                r.fee_earner_name,
                r.created_at.strftime("%Y-%m-%d %H:%M"),
                float(pence_to_pounds_str(r.fees_ex_vat_pence)),
                float(pence_to_pounds_str(r.vat_pence)),
                float(pence_to_pounds_str(r.disbursements_ex_vat_pence)),
            ]
        )
    ws.append(
        [
            "",
            "",
            "",
            "",
            "TOTAL",
            "",
            float(pence_to_pounds_str(totals["fees_ex_vat_pence"])),
            float(pence_to_pounds_str(totals["vat_pence"])),
            float(pence_to_pounds_str(totals["disbursements_ex_vat_pence"])),
        ]
    )
    return _wb_response(wb, "canary-report-billing.xlsx")


def _parse_case_statuses(raw: list[str] | None) -> list[CaseStatus] | None:
    if not raw:
        return None
    out: list[CaseStatus] = []
    for s in raw:
        try:
            out.append(CaseStatus(s))
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid case status: {s!r}",
            ) from exc
    return out


@router.post("/cases")
def post_cases_report(
    body: CasesReportIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    st = _parse_case_statuses(body.statuses)
    rows = report_cases(ids, db, statuses=st)
    if format == "json":
        return {
            "rows": [
                {
                    "case_id": str(r.case_id),
                    "case_number": r.case_number,
                    "client_name": r.client_name,
                    "matter_description": r.matter_description,
                    "status": r.status,
                    "status_label": r.status_label,
                    "fee_earner_name": r.fee_earner_name,
                    "created_at": r.created_at.isoformat(),
                }
                for r in rows
            ]
        }
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    ws.append(["Reference", "Client", "Matter description", "Status", "Fee earner", "Created (UTC)"])
    for r in rows:
        ws.append(
            [
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.status_label,
                r.fee_earner_name,
                r.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )
    return _wb_response(wb, "canary-report-cases.xlsx")


@router.post("/cases-opened")
def post_cases_opened_report(
    body: CasesOpenedReportIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    rows = report_cases_opened(
        ids,
        db,
        date_from=body.date_from,
        date_to=body.date_to,
        include_quote=body.include_quote,
        include_active=body.include_active,
    )
    if format == "json":
        return {
            "rows": [
                {
                    "case_id": str(r.case_id),
                    "case_number": r.case_number,
                    "client_name": r.client_name,
                    "matter_description": r.matter_description,
                    "status": r.status,
                    "status_label": r.status_label,
                    "fee_earner_name": r.fee_earner_name,
                    "source_name": r.source_name or "",
                    "created_at": r.created_at.isoformat(),
                }
                for r in rows
            ]
        }
    wb = Workbook()
    ws = wb.active
    ws.title = "Cases opened"
    ws.append(["Reference", "Client", "Matter description", "Status", "Fee earner", "Source", "Opened (UTC)"])
    for r in rows:
        ws.append(
            [
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.status_label,
                r.fee_earner_name,
                r.source_name or "",
                r.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )
    return _wb_response(wb, "canary-report-cases-opened.xlsx")


@router.post("/events")
def post_events_report(
    body: EventsReportIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    rows = report_events(
        ids,
        db,
        date_from=body.date_from,
        date_to=body.date_to,
        template_ids=body.template_ids,
    )
    if format == "json":
        return {
            "rows": [
                {
                    "event_id": str(r.event_id),
                    "case_number": r.case_number,
                    "matter_description": r.matter_description,
                    "fee_earner_name": r.fee_earner_name,
                    "event_name": r.event_name,
                    "event_date": r.event_date.isoformat() if r.event_date else None,
                    "event_category": r.event_category or "",
                }
                for r in rows
            ]
        }
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    ws.append(
        [
            "Event",
            "Event date",
            "Calendar template",
            "Reference",
            "Matter description",
            "Fee earner",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.event_name,
                r.event_date.isoformat() if r.event_date else "",
                r.event_category or "",
                r.case_number,
                r.matter_description,
                r.fee_earner_name,
            ]
        )
    return _wb_response(wb, "canary-report-events.xlsx")


@router.post("/ledger-activity")
def post_ledger_activity_report(
    body: LedgerActivityReportIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    rows = report_ledger_activity(
        ids,
        db,
        date_from=body.date_from,
        date_to=body.date_to,
        approved_only=body.approved_only,
    )
    if format == "json":
        return {
            "rows": [
                {
                    "pair_id": str(r.pair_id),
                    "case_id": str(r.case_id),
                    "case_number": r.case_number,
                    "client_name": r.client_name or "",
                    "matter_description": r.matter_description,
                    "fee_earner_name": r.fee_earner_name,
                    "posted_at": r.posted_at.isoformat(),
                    "posted_by_name": r.posted_by_name,
                    "description": r.description,
                    "reference": r.reference or "",
                    "amount_pence": r.amount_pence,
                    "client_direction": r.client_direction,
                    "office_direction": r.office_direction,
                    "is_approved": r.is_approved,
                    "contact_label": r.contact_label or "",
                }
                for r in rows
            ]
        }
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger activity"
    ws.append(
        [
            "Posted (UTC)",
            "Reference",
            "Client",
            "Matter",
            "Fee earner",
            "Description",
            "Ledger ref",
            "Amount (£)",
            "Client leg",
            "Office leg",
            "Approved",
            "Posted by",
            "Contact",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.posted_at.strftime("%Y-%m-%d %H:%M"),
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.fee_earner_name,
                r.description,
                r.reference or "",
                pence_to_pounds_str(r.amount_pence),
                r.client_direction or "",
                r.office_direction or "",
                "Yes" if r.is_approved else "Pending",
                r.posted_by_name,
                r.contact_label or "",
            ]
        )
    return _wb_response(wb, "canary-report-ledger-activity.xlsx")


def _serialize_exceptions(report) -> dict:
    def bal_row(r):
        return {
            "case_id": str(r.case_id),
            "case_number": r.case_number,
            "client_name": r.client_name or "",
            "matter_description": r.matter_description,
            "status": r.status,
            "status_label": r.status_label,
            "fee_earner_name": r.fee_earner_name,
            "client_balance_pence": r.client_balance_pence,
            "office_balance_pence": r.office_balance_pence,
        }

    def ledger_row(r):
        return {
            "pair_id": str(r.pair_id),
            "case_id": str(r.case_id),
            "case_number": r.case_number,
            "client_name": r.client_name or "",
            "matter_description": r.matter_description,
            "fee_earner_name": r.fee_earner_name,
            "posted_at": r.posted_at.isoformat(),
            "posted_by_name": r.posted_by_name,
            "description": r.description,
            "amount_pence": r.amount_pence,
            "client_direction": r.client_direction,
            "office_direction": r.office_direction,
            "is_approved": getattr(r, "is_approved", False),
        }

    return {
        "pending_ledger_approvals": [ledger_row(r) for r in report.pending_ledger_approvals],
        "pending_invoices": [
            {
                "invoice_id": str(r.invoice_id),
                "case_id": str(r.case_id),
                "case_number": r.case_number,
                "client_name": r.client_name or "",
                "matter_description": r.matter_description,
                "fee_earner_name": r.fee_earner_name,
                "invoice_number": r.invoice_number,
                "created_at": r.created_at.isoformat(),
                "total_pence": r.total_pence,
            }
            for r in report.pending_invoices
        ],
        "client_balance_closed_archived": [bal_row(r) for r in report.client_balance_closed_archived],
        "negative_client_balance": [bal_row(r) for r in report.negative_client_balance],
        "large_postings": [ledger_row(r) for r in report.large_postings],
    }


@router.post("/aged-debt")
def post_aged_debt_report(
    body: AgedDebtReportIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    rows, bucket_totals = report_aged_debt(ids, db, as_of=body.as_of)
    if format == "json":
        return {
            "rows": [
                {
                    "invoice_id": str(r.invoice_id),
                    "case_id": str(r.case_id),
                    "case_number": r.case_number,
                    "client_name": r.client_name or "",
                    "matter_description": r.matter_description,
                    "fee_earner_name": r.fee_earner_name,
                    "invoice_number": r.invoice_number,
                    "approved_at": r.approved_at.isoformat(),
                    "age_days": r.age_days,
                    "age_bucket": r.age_bucket,
                    "invoice_total_pence": r.invoice_total_pence,
                    "office_balance_pence": r.office_balance_pence,
                }
                for r in rows
            ],
            "bucket_totals_pence": bucket_totals,
        }
    wb = Workbook()
    ws = wb.active
    ws.title = "Aged debt"
    ws.append(
        [
            "Age bucket",
            "Days",
            "Reference",
            "Client",
            "Matter",
            "Fee earner",
            "Invoice",
            "Approved (UTC)",
            "Invoice total (£)",
            "Office balance (£)",
        ]
    )
    for r in rows:
        ws.append(
            [
                r.age_bucket,
                r.age_days,
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.fee_earner_name,
                r.invoice_number,
                r.approved_at.strftime("%Y-%m-%d %H:%M"),
                float(pence_to_pounds_str(r.invoice_total_pence)),
                float(pence_to_pounds_str(r.office_balance_pence)),
            ]
        )
    ws.append([])
    ws.append(["Bucket summary (invoice totals £)", "", "", "", "", "", "", "0-30", "31-60", "61-90", "90+"])
    ws.append(
        [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            float(pence_to_pounds_str(bucket_totals["0-30"])),
            float(pence_to_pounds_str(bucket_totals["31-60"])),
            float(pence_to_pounds_str(bucket_totals["61-90"])),
            float(pence_to_pounds_str(bucket_totals["90+"])),
        ]
    )
    return _wb_response(wb, "canary-report-aged-debt.xlsx")


@router.post("/exceptions")
def post_exceptions_report(
    body: ExceptionsReportIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
    format: Annotated[ReportFormat, Query()] = "json",
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    report = report_exceptions(
        ids,
        db,
        date_from=body.date_from,
        date_to=body.date_to,
        large_posting_min_pence=body.large_posting_min_pence,
    )
    payload = _serialize_exceptions(report)
    if format == "json":
        return payload

    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title: str, headers: list[str], data_rows: list[list]):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for row in data_rows:
            ws.append(row)

    add_sheet(
        "Pending ledger",
        ["Posted (UTC)", "Reference", "Client", "Matter", "Fee earner", "Description", "Amount (£)", "Client leg", "Office leg", "Posted by"],
        [
            [
                r.posted_at.strftime("%Y-%m-%d %H:%M"),
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.fee_earner_name,
                r.description,
                float(pence_to_pounds_str(r.amount_pence)),
                r.client_direction or "",
                r.office_direction or "",
                r.posted_by_name,
            ]
            for r in report.pending_ledger_approvals
        ],
    )
    add_sheet(
        "Pending invoices",
        ["Created (UTC)", "Reference", "Client", "Matter", "Fee earner", "Invoice", "Total (£)"],
        [
            [
                r.created_at.strftime("%Y-%m-%d %H:%M"),
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.fee_earner_name,
                r.invoice_number,
                float(pence_to_pounds_str(r.total_pence)),
            ]
            for r in report.pending_invoices
        ],
    )
    add_sheet(
        "Closed archived client bal",
        ["Reference", "Client", "Matter", "Status", "Fee earner", "Client balance (£)", "Office balance (£)"],
        [
            [
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.status_label,
                r.fee_earner_name,
                float(pence_to_pounds_str(r.client_balance_pence)),
                float(pence_to_pounds_str(r.office_balance_pence)),
            ]
            for r in report.client_balance_closed_archived
        ],
    )
    add_sheet(
        "Negative client balance",
        ["Reference", "Client", "Matter", "Status", "Fee earner", "Client balance (£)", "Office balance (£)"],
        [
            [
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.status_label,
                r.fee_earner_name,
                float(pence_to_pounds_str(r.client_balance_pence)),
                float(pence_to_pounds_str(r.office_balance_pence)),
            ]
            for r in report.negative_client_balance
        ],
    )
    add_sheet(
        "Large postings",
        ["Posted (UTC)", "Reference", "Client", "Matter", "Fee earner", "Description", "Amount (£)", "Approved", "Posted by"],
        [
            [
                r.posted_at.strftime("%Y-%m-%d %H:%M"),
                r.case_number,
                r.client_name or "",
                r.matter_description,
                r.fee_earner_name,
                r.description,
                float(pence_to_pounds_str(r.amount_pence)),
                "Yes" if r.is_approved else "Pending",
                r.posted_by_name,
            ]
            for r in report.large_postings
        ],
    )
    return _wb_response(wb, "canary-report-exceptions.xlsx")


def _pack_params(body: AccountantPackIn) -> dict:
    return {
        "period_end_date": body.period_end_date,
        "date_from": body.date_from,
        "date_to": body.date_to,
        "include_balances": body.include_balances,
        "include_billing": body.include_billing,
        "include_ledger_activity": body.include_ledger_activity,
        "include_aged_debt": body.include_aged_debt,
        "include_exceptions": body.include_exceptions,
        "include_reconcile_doc": body.include_reconcile_doc,
        "large_posting_min_pence": body.large_posting_min_pence,
    }


def _preview_to_out(preview) -> AccountantPackPreviewOut:
    return AccountantPackPreviewOut(
        period_end_date=preview.period_end_date,
        activity_date_from=preview.activity_date_from,
        activity_date_to=preview.activity_date_to,
        fee_earner_count=preview.fee_earner_count,
        reconcile_doc_available=preview.reconcile_doc_available,
        sections=[
            AccountantPackSectionOut(
                key=s.key,
                label=s.label,
                included=s.included,
                row_count=s.row_count,
                note=s.note,
            )
            for s in preview.sections
        ],
    )


@router.post("/accountant-pack/preview", response_model=AccountantPackPreviewOut)
def post_accountant_pack_preview(
    body: AccountantPackIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> AccountantPackPreviewOut:
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    if not any(
        [
            body.include_balances,
            body.include_billing,
            body.include_ledger_activity,
            body.include_aged_debt,
            body.include_exceptions,
            body.include_reconcile_doc,
        ]
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Select at least one section to include in the export pack.",
        )
    preview = preview_accountant_pack(db, fee_earner_user_ids=ids, **_pack_params(body))
    return _preview_to_out(preview)


@router.post("/accountant-pack")
def post_accountant_pack(
    body: AccountantPackIn,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    ids = enforce_fee_earner_ids(user, db, body.fee_earner_user_ids)
    if not any(
        [
            body.include_balances,
            body.include_billing,
            body.include_ledger_activity,
            body.include_aged_debt,
            body.include_exceptions,
            body.include_reconcile_doc,
        ]
    ):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Select at least one section to include in the export pack.",
        )
    generated_by = (user.display_name or user.email or "").strip() or str(user.id)
    result = build_accountant_pack(
        db,
        fee_earner_user_ids=ids,
        generated_by_name=generated_by,
        **_pack_params(body),
    )
    bio = BytesIO(result.zip_bytes)
    return StreamingResponse(
        bio,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{result.filename}"'},
    )
