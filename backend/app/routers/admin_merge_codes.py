"""Admin merge code catalog: list, bulk edit, Excel export/import."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.audit import log_event
from app.db import get_db
from app.deps import require_admin
from app.models import MergeCodeCatalog, User
from app.schemas import MergeCodeCatalogBulkUpdate, MergeCodeCatalogImportResult, MergeCodeCatalogOut

router = APIRouter(prefix="/admin/merge-codes", tags=["admin-merge-codes"])


@router.get("/export.xlsx")
def export_merge_codes_xlsx(admin: User = Depends(require_admin), db: Session = Depends(get_db)) -> StreamingResponse:
    rows = (
        db.execute(select(MergeCodeCatalog).order_by(MergeCodeCatalog.sort_order.asc(), MergeCodeCatalog.code.asc()))
        .scalars()
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Merge codes"
    ws.append(["Code", "Description"])
    for r in rows:
        ws.append([r.code, r.description])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    log_event(
        db,
        actor_user_id=admin.id,
        action="merge_code_catalog.export",
        entity_type="merge_code_catalog",
        entity_id="export.xlsx",
        meta={"rows": len(rows)},
    )
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="canary-merge-codes.xlsx"'},
    )


@router.get("", response_model=list[MergeCodeCatalogOut])
def list_merge_codes(admin: User = Depends(require_admin), db: Session = Depends(get_db)) -> list[MergeCodeCatalogOut]:
    rows = (
        db.execute(select(MergeCodeCatalog).order_by(MergeCodeCatalog.sort_order.asc(), MergeCodeCatalog.code.asc()))
        .scalars()
        .all()
    )
    return [MergeCodeCatalogOut.model_validate(r, from_attributes=True) for r in rows]


@router.patch("", response_model=list[MergeCodeCatalogOut])
def bulk_update_merge_codes(
    payload: MergeCodeCatalogBulkUpdate,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[MergeCodeCatalogOut]:
    if not payload.items:
        rows = (
            db.execute(select(MergeCodeCatalog).order_by(MergeCodeCatalog.sort_order.asc(), MergeCodeCatalog.code.asc()))
            .scalars()
            .all()
        )
        return [MergeCodeCatalogOut.model_validate(r, from_attributes=True) for r in rows]

    if len(payload.items) > 5000:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Too many rows in one request.")

    for item in payload.items:
        row = db.get(MergeCodeCatalog, item.code.strip())
        if row is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unknown merge code: {item.code!r}. Add new codes in software, then restart to sync.",
            )
        row.description = item.description.strip()
        row.updated_at = datetime.utcnow()
        db.add(row)

    db.commit()

    rows = (
        db.execute(select(MergeCodeCatalog).order_by(MergeCodeCatalog.sort_order.asc(), MergeCodeCatalog.code.asc()))
        .scalars()
        .all()
    )
    log_event(
        db,
        actor_user_id=admin.id,
        action="merge_code_catalog.bulk_update",
        entity_type="merge_code_catalog",
        entity_id="*",
        meta={"count": len(payload.items)},
    )
    return [MergeCodeCatalogOut.model_validate(r, from_attributes=True) for r in rows]


@router.post("/import", response_model=MergeCodeCatalogImportResult)
async def import_merge_codes_xlsx(
    upload: UploadFile = File(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> MergeCodeCatalogImportResult:
    raw = await upload.read()
    if not raw or len(raw) > 12 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Upload a spreadsheet under 12 MB.")

    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not read Excel file: {exc}",
        ) from exc

    ws = wb.active
    updated = 0
    skipped_unknown = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i > 20000:
            break
        if not row:
            continue
        code_raw = row[0]
        if code_raw is None or str(code_raw).strip() == "":
            continue
        code = str(code_raw).strip()
        desc_cell = row[1] if len(row) > 1 else None
        desc = "" if desc_cell is None else str(desc_cell)

        cat = db.get(MergeCodeCatalog, code)
        if cat is None:
            skipped_unknown += 1
            continue
        cat.description = desc.strip()
        cat.updated_at = datetime.utcnow()
        db.add(cat)
        updated += 1

    db.commit()
    log_event(
        db,
        actor_user_id=admin.id,
        action="merge_code_catalog.import",
        entity_type="merge_code_catalog",
        entity_id="import.xlsx",
        meta={"updated": updated, "skipped_unknown": skipped_unknown},
    )
    return MergeCodeCatalogImportResult(updated=updated, skipped_unknown=skipped_unknown)
