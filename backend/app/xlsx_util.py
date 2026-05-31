"""Spreadsheet (.xlsx) helpers for quote / fee-scale compose."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

_MERGE_TOKEN_RE = re.compile(r"\[(?:[biu]+:)?([A-Z0-9_]+)\]", re.IGNORECASE)


@dataclass
class XlsxGrid:
    """Rectangular fee-table data extracted from a worksheet for Word table insertion."""

    rows: list[list[str]]
    merges: list[tuple[int, int, int, int]] = field(default_factory=list)  # r0,c0,r1,c1 inclusive
    bold: set[tuple[int, int]] = field(default_factory=set)


def write_blank_xlsx_bytes() -> bytes:
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Quote"
    wb.save(buf)
    return buf.getvalue()


def _replace_merge_tokens_in_text(text: str, fields: Mapping[str, str]) -> str:
    if not text or "[" not in text:
        return text

    def repl(m: re.Match[str]) -> str:
        key = f"[{m.group(1).upper()}]"
        return fields.get(key, m.group(0))

    return _MERGE_TOKEN_RE.sub(repl, text)


def merge_precedent_codes_in_xlsx_bytes(src_bytes: bytes, fields: Mapping[str, str]) -> bytes:
    """Replace ``[CODE]`` placeholders in all worksheet cells."""
    wb = load_workbook(io.BytesIO(src_bytes))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and "[" in val:
                    cell.value = _replace_merge_tokens_in_text(val, fields)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def validate_xlsx_package_bytes(data: bytes) -> None:
    """Raise ValueError if ``data`` is not a readable .xlsx workbook."""
    try:
        load_workbook(io.BytesIO(data), read_only=True).close()
    except Exception as exc:
        raise ValueError(f"Not a valid .xlsx workbook: {exc}") from exc


def _cell_display(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def extract_xlsx_grid(xlsx_bytes: bytes, *, sheet_index: int = 0) -> XlsxGrid:
    """Read the used range of a worksheet as a grid suitable for a Word table."""
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    if sheet_index >= len(wb.worksheets):
        return XlsxGrid(rows=[[" "]])
    ws = wb.worksheets[sheet_index]

    min_r, min_c, max_r, max_c = 10**9, 10**9, 0, 0

    def include(r: int, c: int) -> None:
        nonlocal min_r, min_c, max_r, max_c
        min_r = min(min_r, r)
        min_c = min(min_c, c)
        max_r = max(max_r, r)
        max_c = max(max_c, c)

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None and str(cell.value).strip():
                include(cell.row, cell.column)
    for mr in ws.merged_cells.ranges:
        include(mr.min_row, mr.min_col)
        include(mr.max_row, mr.max_col)

    if max_r == 0:
        return XlsxGrid(rows=[[" "]])

    nrows = max_r - min_r + 1
    ncols = max_c - min_c + 1
    grid: list[list[str]] = [[""] * ncols for _ in range(nrows)]
    bold: set[tuple[int, int]] = set()

    for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            r_i = cell.row - min_r
            c_i = cell.column - min_c
            grid[r_i][c_i] = _cell_display(cell.value)
            if cell.font and cell.font.bold:
                bold.add((r_i, c_i))

    merges: list[tuple[int, int, int, int]] = []
    for mr in ws.merged_cells.ranges:
        if mr.max_row < min_r or mr.min_row > max_r or mr.max_col < min_c or mr.min_col > max_c:
            continue
        merges.append(
            (
                mr.min_row - min_r,
                mr.min_col - min_c,
                mr.max_row - min_r,
                mr.max_col - min_c,
            )
        )

    return XlsxGrid(rows=grid, merges=merges, bold=bold)
