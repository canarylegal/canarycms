from io import BytesIO

from openpyxl import Workbook

from app.xlsx_util import autofit_workbook, autofit_worksheet_columns


def test_autofit_worksheet_columns_sets_width_from_content() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Reference", "Client"])
    ws.append(["TST/001", "A short name"])
    ws.append(["TST/002", "A much longer client name here"])

    autofit_worksheet_columns(ws)

    assert ws.column_dimensions["A"].width >= 8
    assert ws.column_dimensions["B"].width > ws.column_dimensions["A"].width


def test_autofit_workbook_applies_to_all_sheets() -> None:
    wb = Workbook()
    wb.active.append(["Only"])
    wb.create_sheet("Second").append(["Second sheet with a longer header"])

    autofit_workbook(wb)

    assert wb.worksheets[0].column_dimensions["A"].width >= 8
    assert wb.worksheets[1].column_dimensions["A"].width > wb.worksheets[0].column_dimensions["A"].width


def test_autofit_workbook_round_trips_through_save() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Description"])
    ws.append(["Payment on account for conveyancing fees"])

    autofit_workbook(wb)
    bio = BytesIO()
    wb.save(bio)
    assert bio.getvalue()
